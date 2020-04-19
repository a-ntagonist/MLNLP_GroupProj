import csv
import io
import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from statsmodels.sandbox.regression.predstd import wls_prediction_std

def readkeys(fn):
    f = open(fn+'.txt').read().split()
    return f

def getstats(lang, stat):
    stats = dict()
    wb = load_workbook('..\\data\\{}\\{}_stats.xlsx'.format(lang, stat)) #statistics for key concept
    ws = wb.active
    for row in ws.rows:
        stats[row[0].value] = row[1].value # {'key concept': stat(float)} dictionary
    ans = dict()
    wb = load_workbook('..\\data\\{}\\{}_words.xlsx'.format(lang, stat)) #lemmas for key concepts
    ws = wb.active
    for row in ws.rows:
        if len(row)>1:
            words = tuple(filter(lambda x:x!=None, map(lambda x:x.value, row[1:])))
            ans[(row[0].value, words)] = stats[row[0].value] #{('key concept', [lemmas*]):stat(float)} dictionary
    return ans


class wordvec:
    def __init__(self, lang, fname, need):
        fin = io.open('..\\data\\{}\\'.format(lang)+fname, 'r', encoding='utf-8', newline='\n', errors='ignore')
        n, d = map(int, fin.readline().split())
        data = {}
        for line in fin:
            tokens = line.rstrip().split(' ')
            if tokens[0] in need:
                data[tokens[0]] = list(map(float, tokens[1:]))
        self.vecdict = data# dictionary token:nparray

    def word(self, word): # return nparray or None for OOV
        if word in self.vecdict:
            return self.vecdict[word]
        else:
            return None

def avgvec(words, vectors):
    K = list(filter(lambda x:x!=None, list(map(vectors.word, words))))
    if len(K) == 0:
        print('None of {} are in word vector!'.format(words))
        return np.array([None for i in range(300)])
    K = np.mean(K, axis=0)
    #print(len(K), K[0], type(K))
    return K


def getcossim(vecA, vecB):
    return np.linalg.norm(vecA-vecB)

if __name__ == '__main__':

    lang = 'japanese' #or 'japanese'
    statname = 'adjective' #or 'participation' or 'mturk'

    group_A = readkeys('..\\data\\{}\\female_keys'.format(lang)) #list [w1, ... wD]
    group_B = readkeys('..\\data\\{}\\male_keys'.format(lang)) #list [w1, ... wD]
    need = set(group_A).union(set(group_B)) #words we need vectors for
    alljobs = getstats(lang, statname) #dict{(keyword, (lemmas)):stat as float} dictionary
    for jobs, stat in alljobs.items():
        need = need.union(set(jobs[1])) #we need vectors for all lemmas

    vectors = wordvec(lang, 'cc.{}.300.vec'.format(lang[:2]), need) #maps words to vectors, with only needed words to save memory
    groupvec_A = avgvec(group_A, vectors) #compute group vector for group A
    groupvec_B = avgvec(group_B, vectors) #compute group vector for group A
    c = 0 #for counting OOV keywords
    results = []
    for jobs, stat in alljobs.items():
        jobvec = avgvec(jobs[1], vectors) #compute average of vectors for lemmas
        if jobvec.any() != None: #if the average vector is valid without nan value,
            bias = getcossim(groupvec_B, jobvec)-getcossim(groupvec_A, jobvec) #get difference distance to each group vector
            results.append([jobs[0], bias, stat]) #record bias score for job and actual statistic
        else:
            c+=1# count if the average vector is invalid (none of the lemmas had matching word vectors)
    print('Total number of key concepts:', len(alljobs))
    print('{} OOV concepts, {:.3f}% of total concepts'.format(c, c/len(alljobs)))
    results = pd.DataFrame(results, columns=['job', 'bias', 'stat'])
    results.to_csv('..\\Results\\{}_{}_results.csv'.format(lang, statname), encoding='utf-8') #save results as csv file


    #run linear regression for bias score vs statistics
    X = sm.add_constant(results.bias.astype(float))
    Y = results.stat.astype(float)
    model = sm.OLS(Y, X).fit()
    print('p value: ', model.f_pvalue)
    print('r squared: ', model.rsquared)
    print(model.params)

    with open('..\\Results\\{}_{}_results.txt'.format(lang, statname), 'w') as f:
        f.write('Total number of key concepts: {}\n'.format(len(alljobs)))
        f.write('{} OOV concepts, {:.3f}% of total concepts\n'.format(c, c/len(alljobs)))
        f.write('p value: {}\n'.format(model.f_pvalue))
        f.write('r squared: {}\n'.format(model.rsquared))
        f.write(str(model.params))



    #plot result
    plt.rcParams["figure.figsize"] = (20,10)
    x = np.linspace(-0.5, 0.5)
    conf = model.conf_int(0.05) #plotting confidence range
    plt.fill_between(x, conf[0][0]+conf[0][1]*x, conf[1][0]+conf[0][1]*x, color=(0.75, 0.75, 0.75, 1))
    plt.fill_between(x, conf[0][0]+conf[1][1]*x, conf[1][0]+conf[1][1]*x, color=(0.75, 0.75, 0.75, 1))
    plt.fill_between(x, conf[0][0]+conf[0][1]*x, conf[1][0]+conf[1][1]*x, color=(0.75, 0.75, 0.75, 1))
    plt.scatter(results.bias, results.stat, color=(0, 0.5, 0.5, 1)) #plot datapoints
    plt.plot(x, model.params[0]+model.params[1]*x, 'r-') #plot model
    plt.grid()
    plt.xlabel('Embedding Bias Score', fontsize=40)
    plt.ylabel('Statistics', fontsize=40)
    plt.xticks(fontsize=25)
    plt.yticks(fontsize=25)
    plt.xlim(-0.5, 0.5)
    plt.savefig('..\\Results\\{}_{}_results.png'.format(lang, statname))
