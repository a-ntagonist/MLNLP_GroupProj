import csv
import io
import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl import load_workbook
import matplotlib.pyplot as plt

from statsmodels.sandbox.regression.predstd import wls_prediction_std

def readkeys(fn):
    f = open('data\\'+fn+'.txt').read().split()
    return f

def getstats():
    stats = dict()
    wb = load_workbook('data\\adj.xlsx')
    ws = wb.active
    for row in ws.rows:
        stats[row[0].value] = row[1].value
    ans = dict()
    wb = load_workbook('data\\adj_words.xlsx')
    ws = wb.active
    for row in ws.rows:
        if len(row)>1:
            words = tuple(filter(lambda x:x!=None, map(lambda x:x.value, row[1:])))
            ans[(row[0].value, words)] = stats[row[0].value]
    return ans


class wordvec:
    def __init__(self, fname, need):
        fin = io.open('data\\'+fname, 'r', encoding='utf-8', newline='\n', errors='ignore')
        n, d = map(int, fin.readline().split())
        data = {}
        for line in fin:
            tokens = line.rstrip().split(' ')
            if tokens[0] in need:
                data[tokens[0]] = list(map(float, tokens[1:]))
        self.vecdict = data# dictionary token:nparray

    def word(self, word): # return nparray or None for OOV
        if word in self.vecdict:
            return self.vecdict[word]
        else:
            return None

def avgvec(words, vectors):
    K = list(filter(lambda x:x!=None, list(map(vectors.word, words))))
    if len(K) == 0:
        print('None of {} are in word vector!'.format(words))
        return np.array([None for i in range(300)])
    K = np.mean(K, axis=0)
    #print(len(K), K[0], type(K))
    return K


def getcossim(vecA, vecB):
    return np.linalg.norm(vecA-vecB)

if __name__ == '__main__':
    group_A = readkeys('female_pairs') #list [w1, ... wD]
    group_B = readkeys('male_pairs') #list [w1, ... wD]
    need = set(group_A).union(set(group_B))
    alljobs = getstats() #dict{(keyword, (words)):float(0~1)}
    for jobs, stat in alljobs.items():
        need = need.union(set(jobs[1]))

    vectors = wordvec('cc.ko.300.vec', need) # wordvec object
    groupvec_A = avgvec(group_A, vectors)
    groupvec_B = avgvec(group_B, vectors)
    c = 0
    results = []
    for jobs, stat in alljobs.items():
        jobvec = avgvec(jobs[1], vectors)
        if jobvec.any() != None:
            bias = getcossim(groupvec_B, jobvec)-getcossim(groupvec_A, jobvec)
            results.append([jobs[0], bias, stat])
        else:
            c+=1
    print(len(alljobs))
    print(c, c/len(alljobs))
    results = pd.DataFrame(results, columns=['job', 'bias', 'stat'])
    results.to_csv('results.csv', encoding='utf-8')

    X = sm.add_constant(results.bias.astype(float))
    Y = results.stat.astype(float)

    model = sm.OLS(Y, X).fit()
    x = np.linspace(-0.5, 0.5)
    print('p value: ', model.f_pvalue)
    print('r squared: ', model.rsquared)
    print(model.params)
    plt.scatter(results.bias, results.stat)
    plt.plot(x, model.params[0]+model.params[1]*x, 'r-')
    conf = model.conf_int(0.05)
    print(conf[0][0])
    plt.fill_between(x, conf[0][0]+conf[0][1]*x, conf[1][0]+conf[0][1]*x, color=(0.75, 0.75, 0.75, 1))
    plt.fill_between(x, conf[0][0]+conf[1][1]*x, conf[1][0]+conf[1][1]*x, color=(0.75, 0.75, 0.75, 1))
    plt.fill_between(x, conf[0][0]+conf[0][1]*x, conf[1][0]+conf[1][1]*x, color=(0.75, 0.75, 0.75, 1))

    plt.scatter(results.bias, results.stat, color=(0, 0.5, 0.5, 1))
    plt.plot(x, model.params[0]+model.params[1]*x, 'r-')
    plt.grid()
    plt.xlabel('Embedding Bias Score', fontsize=40)
    plt.ylabel('Human Labeled Bias Score', fontsize=40)
    plt.xticks(fontsize=25)
    plt.yticks(fontsize=25)
    plt.xlim(-0.5, 0.5)
    plt.show()
